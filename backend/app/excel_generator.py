"""Excel generation with openpyxl"""

from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import ShoppingItem, WindowData

HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BLUE_HEADER_FILL = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
BLUE_HEADER_FONT = Font(bold=True, color="FFFFFF")


def generate_window_excel(window_data: WindowData, output_path: str) -> str:
    """Generate professional Excel workbook for a single window"""

    workbook = Workbook()
    workbook.remove(workbook.active)

    create_overview_sheet(workbook, window_data)
    create_precut_sheet(workbook, window_data)
    create_cut_sheet(workbook, window_data)
    create_shopping_sheet(workbook, window_data)
    create_glazing_sheet(workbook, window_data)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return str(path)


def create_overview_sheet(workbook: Workbook, window_data: WindowData) -> None:
    sheet = workbook.create_sheet("Overview", 0)

    sheet['A1'] = 'SASH WINDOW SPECIFICATION'
    sheet['A1'].font = Font(size=16, bold=True)
    sheet.merge_cells('A1:D1')

    info_rows = [
        ("Date:", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Configuration:", window_data.config),
        ("Description:", window_data.glazing.configuration),
        ("", ""),
        ("Frame Width:", f"{window_data.frame.width} mm"),
        ("Frame Height:", f"{window_data.frame.height} mm"),
        ("Sash Width:", f"{window_data.sash.width} mm"),
        ("Sash Height:", f"{window_data.sash.height} mm"),
    ]

    row_index = 3
    for label, value in info_rows:
        sheet[f'A{row_index}'] = label
        sheet[f'B{row_index}'] = value
        sheet[f'A{row_index}'].font = Font(bold=True)
        row_index += 1

    sheet.column_dimensions['A'].width = 22
    sheet.column_dimensions['B'].width = 26


def create_precut_sheet(workbook: Workbook, window_data: WindowData) -> None:
    sheet = workbook.create_sheet("Pre-cut List")
    headers = ['Element', 'Width (mm)', 'Pre-cut Length (mm)', 'Quantity', 'Material']

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    row = 2
    for component in iter_components(window_data.components):
        sheet.cell(row=row, column=1, value=component.get('element'))
        sheet.cell(row=row, column=2, value=component.get('width'))
        sheet.cell(row=row, column=3, value=component.get('preCutLength') or component.get('length'))
        sheet.cell(row=row, column=4, value=component.get('quantity', 1))
        sheet.cell(row=row, column=5, value=component.get('material'))
        row += 1

    autosize_columns(sheet, len(headers))


def create_cut_sheet(workbook: Workbook, window_data: WindowData) -> None:
    sheet = workbook.create_sheet("Cut List")
    headers = ['Element', 'Cut Length (mm)', 'Quantity', 'Material']

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    row = 2
    for component in iter_components(window_data.components):
        sheet.cell(row=row, column=1, value=component.get('element'))
        sheet.cell(row=row, column=2, value=component.get('cutLength') or component.get('length'))
        sheet.cell(row=row, column=3, value=component.get('quantity', 1))
        sheet.cell(row=row, column=4, value=component.get('material'))
        row += 1

    autosize_columns(sheet, len(headers))


def create_shopping_sheet(workbook: Workbook, window_data: WindowData) -> None:
    sheet = workbook.create_sheet("Shopping List")
    headers = ['Material', 'Specification', 'Quantity', 'Unit']

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    shopping_items = gather_shopping_items(window_data)
    row = 2
    for item in shopping_items:
        sheet.cell(row=row, column=1, value=item.material)
        sheet.cell(row=row, column=2, value=item.specification)
        sheet.cell(row=row, column=3, value=item.quantity)
        sheet.cell(row=row, column=4, value=item.unit)
        row += 1

    autosize_columns(sheet, len(headers))


def create_glazing_sheet(workbook: Workbook, window_data: WindowData) -> None:
    sheet = workbook.create_sheet("Glazing")
    headers = ['Pane #', 'Width (mm)', 'Height (mm)', 'Position', 'Type']

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = BLUE_HEADER_FONT
        cell.fill = BLUE_HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    glazing_type = window_data.options.get('glazingType', '4mm Clear')
    row = 2
    for pane in window_data.glazing.panes:
        sheet.cell(row=row, column=1, value=f"#{pane.id}")
        sheet.cell(row=row, column=2, value=round(pane.width, 1))
        sheet.cell(row=row, column=3, value=round(pane.height, 1))
        sheet.cell(row=row, column=4, value=pane.position)
        sheet.cell(row=row, column=5, value=glazing_type)
        row += 1

    autosize_columns(sheet, len(headers))


def iter_components(components: Dict) -> Iterable[Dict]:
    for group in components.values():
        if isinstance(group, dict):
            for component in group.values():
                if isinstance(component, dict) and component.get('element'):
                    yield component
        elif isinstance(group, list):
            for component in group:
                if isinstance(component, dict) and component.get('element'):
                    yield component


def gather_shopping_items(window_data: WindowData) -> List[ShoppingItem]:
    items: List[ShoppingItem] = []
    if window_data.shopping:
        items.extend(window_data.shopping)
    else:
        for group in window_data.components.values():
            if isinstance(group, list):
                for item in group:
                    if isinstance(item, ShoppingItem):
                        items.append(item)
    return items


def autosize_columns(sheet, column_count: int) -> None:
    for col in range(1, column_count + 1):
        column = get_column_letter(col)
        max_length = 0
        for cell in sheet[column]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column].width = max(12, max_length + 2)
