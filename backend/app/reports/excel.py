"""Excel report generation using openpyxl."""
from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import OptimizationResponse, ProjectRead, ReportRequest
from .pdf import _decode_image

HEADER_FILL = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CENTER = Alignment(horizontal="center", vertical="center")


def _set_column_widths(sheet, widths: Dict[int, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width


def _write_header(sheet, row: int, titles: Iterable[str]) -> None:
    for column, title in enumerate(titles, start=1):
        cell = sheet.cell(row=row, column=column, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def build_workbook(report: ReportRequest, output_path: str) -> Path:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"

    project: ProjectRead = report.project
    overview["A1"] = "Project"
    overview["B1"] = project.name
    overview["A2"] = "Client"
    overview["B2"] = project.client or "—"
    overview["A3"] = "Material"
    overview["B3"] = project.material or "—"
    overview["A4"] = "Section"
    overview["B4"] = project.section_sizes or "—"

    # Cut List sheet
    cut_sheet = workbook.create_sheet("Cut List")
    _write_header(cut_sheet, 1, ["ID", "Type", "Section", "Material", "Length", "Qty", "Drawing"])
    _set_column_widths(cut_sheet, {1: 20, 2: 18, 3: 18, 4: 18, 5: 12, 6: 10, 7: 18})

    for idx, component in enumerate(project.payload.components, start=2):
        cut_sheet.cell(row=idx, column=1, value=component.id)
        cut_sheet.cell(row=idx, column=2, value=component.type)
        cut_sheet.cell(row=idx, column=3, value=component.section)
        cut_sheet.cell(row=idx, column=4, value=component.material)
        cut_sheet.cell(row=idx, column=5, value=component.length)
        cut_sheet.cell(row=idx, column=6, value=component.quantity)
        if component.drawing:
            img = Image(_decode_image(component.drawing))
            img.width = 120
            img.height = 50
            cut_sheet.add_image(img, f"{get_column_letter(7)}{idx}")

    # Pre-pre cut sheet
    if report.optimization:
        pre_sheet = workbook.create_sheet("Pre-Pre Cut")
        _write_header(pre_sheet, 1, ["Bar", "Section", "Material", "Cuts", "Waste", "Utilization %"])
        _set_column_widths(pre_sheet, {1: 18, 2: 18, 3: 18, 4: 28, 5: 12, 6: 16})
        row = 2
        for group in report.optimization.groups:
            for bar in group.bars:
                pre_sheet.cell(row=row, column=1, value=bar.barId)
                pre_sheet.cell(row=row, column=2, value=group.section)
                pre_sheet.cell(row=row, column=3, value=group.material)
                pre_sheet.cell(row=row, column=4, value=", ".join(str(cut) for cut in bar.cuts))
                pre_sheet.cell(row=row, column=5, value=bar.waste)
                pre_sheet.cell(row=row, column=6, value=bar.utilization * 100)
                row += 1

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output))
    return output
